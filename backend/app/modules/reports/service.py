import uuid
import io
from decimal import Decimal
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi import HTTPException

from .queries.trial_balance import get_trial_balance_data
from .queries.general_ledger import get_ledger_entries
from .queries.stock_summary import get_stock_summary_data
from .queries.financial_statements import get_group_balances
from .queries.warehouse_reports import get_warehouse_stock, get_transfer_history
from .schemas.trial_balance import TrialBalanceItem, TrialBalanceResponse
from .schemas.general_ledger import LedgerEntryItem, GeneralLedgerResponse
from .schemas.stock_summary import StockSummaryItem, StockSummaryResponse
from .schemas.dashboard import DashboardMetrics
from .schemas.financial_statements import ProfitLossResponse, BalanceSheetResponse, FinancialStatementItem, BalanceSheetSection
from .schemas.warehouse_reports import WarehouseStockResponse, WarehouseStockItem, TransferReportItem
from app.modules.masters.models import Ledger, Warehouse


class ReportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def get_trial_balance(self, company_id: uuid.UUID, fy_id: uuid.UUID) -> TrialBalanceResponse:
        data = await get_trial_balance_data(self.db, company_id, fy_id)

        items = []
        grand_total_debit = Decimal("0.00")
        grand_total_credit = Decimal("0.00")

        for row in data:
            opening = Decimal(str(row.opening_balance))
            # Adjust opening based on type
            if row.opening_balance_type == "CREDIT":
                opening = -opening

            debit = Decimal(str(row.debit_total))
            credit = Decimal(str(row.credit_total))
            closing = opening + debit - credit

            items.append(TrialBalanceItem(
                ledger_id=row.ledger_id,
                ledger_name=row.ledger_name,
                opening_balance=float(opening),
                debit_total=float(debit),
                credit_total=float(credit),
                closing_balance=float(closing)
            ))
            grand_total_debit += debit
            grand_total_credit += credit

        return TrialBalanceResponse(
            items=items,
            total_debit=float(grand_total_debit),
            total_credit=float(grand_total_credit),
            is_balanced=abs(grand_total_debit - grand_total_credit) < 0.01
        )

    async def get_general_ledger(self, company_id: uuid.UUID, fy_id: uuid.UUID, ledger_id: uuid.UUID) -> GeneralLedgerResponse:
        ledger = await self.db.get(Ledger, ledger_id)
        if not ledger or ledger.company_id != company_id:
            raise HTTPException(status_code=404, detail="Ledger not found")

        opening = Decimal(str(ledger.opening_balance))
        if ledger.opening_balance_type == "CREDIT":
            opening = -opening

        rows = await get_ledger_entries(self.db, company_id, fy_id, ledger_id)

        entries = []
        current_bal = opening
        for row in rows:
            dr = Decimal(str(row.debit))
            cr = Decimal(str(row.credit))
            current_bal += (dr - cr)
            entries.append(LedgerEntryItem(
                date=row.date,
                voucher_number=row.voucher_number,
                voucher_type=row.voucher_type,
                narration=row.narration,
                debit=float(dr),
                credit=float(cr),
                running_balance=float(current_bal)
            ))

        return GeneralLedgerResponse(
            ledger_id=ledger.id,
            ledger_name=ledger.name,
            opening_balance=float(opening),
            entries=entries,
            closing_balance=float(current_bal)
        )

    async def get_stock_summary(self, company_id: uuid.UUID, fy_id: uuid.UUID) -> StockSummaryResponse:
        data = await get_stock_summary_data(self.db, company_id, fy_id)

        items = []
        total_value = Decimal("0.00")

        for row in data:
            opening_qty = float(row.opening_qty)
            inward = float(row.inward_qty)
            outward = float(row.outward_qty)
            closing = opening_qty + inward - outward

            # Average cost logic: (Opening Value + Inward Value) / (Opening Qty + Inward Qty)
            # Actually, WAC is computed during posting and cached.
            # But si.average_cost is the LATEST company wide cost.
            # For reporting, we use the cached average_cost from stock_items OR compute it.
            # Let's use the row's average cost (which might need to be adjusted for period).

            # Simple approach: use si.average_cost
            avg_cost = Decimal(str(row.opening_avg_cost))
            value = Decimal(str(closing)) * avg_cost

            items.append(StockSummaryItem(
                item_id=row.item_id,
                item_name=row.item_name,
                opening_qty=opening_qty,
                inward_qty=inward,
                outward_qty=outward,
                closing_qty=closing,
                average_cost=float(avg_cost),
                stock_value=float(value)
            ))
            total_value += value

        return StockSummaryResponse(
            items=items,
            total_value=float(total_value)
        )

    async def get_dashboard_metrics(self, company_id: uuid.UUID, fy_id: uuid.UUID) -> DashboardMetrics:
        # Implementation of dashboard metrics using existing report logic
        # 1. Total Sales from Trial Balance (Income nature)
        # 2. Total Purchases from Trial Balance (Expense nature)
        # 3. Total Receivables (Sundry Debtors)
        # 4. Total Payables (Sundry Creditors)
        # 5. Inventory Value from Stock Summary

        tb = await self.get_trial_balance(company_id, fy_id)
        stock = await self.get_stock_summary(company_id, fy_id)

        # This is simplified. In a real ERP, we filter by account group name or nature.
        # Since we seeded groups:
        sales = sum(item.credit_total for item in tb.items if "Sales" in item.ledger_name)
        purchases = sum(item.debit_total for item in tb.items if "Purchase" in item.ledger_name)

        # Receivables usually have Debit balance
        receivables = sum(item.closing_balance for item in tb.items if item.closing_balance > 0)
        # This is very rough, should ideally check if parent group is Sundry Debtors

        return DashboardMetrics(
            total_sales=float(sales),
            total_purchases=float(purchases),
            total_receivables=float(receivables),
            total_payables=0.0, # Placeholder
            inventory_value=stock.total_value
        )

    async def get_profit_loss(self, company_id: uuid.UUID, fy_id: uuid.UUID) -> ProfitLossResponse:
        data = await get_group_balances(self.db, company_id, fy_id)

        income_items = []
        expense_items = []
        total_income = Decimal("0.00")
        total_expense = Decimal("0.00")

        for row in data:
            bal = Decimal(str(row.balance))
            if row.nature == "INCOME":
                # Income usually has credit balance (negative in our Dr-Cr math)
                abs_bal = -bal
                income_items.append(FinancialStatementItem(name=row.group_name, amount=float(abs_bal)))
                total_income += abs_bal
            elif row.nature == "EXPENSE":
                abs_bal = bal
                expense_items.append(FinancialStatementItem(name=row.group_name, amount=float(abs_bal)))
                total_expense += abs_bal

        return ProfitLossResponse(
            income=income_items,
            expenses=expense_items,
            total_income=float(total_income),
            total_expenses=float(total_expense),
            net_profit=float(total_income - total_expense)
        )

    async def get_balance_sheet(self, company_id: uuid.UUID, fy_id: uuid.UUID) -> BalanceSheetResponse:
        data = await get_group_balances(self.db, company_id, fy_id)

        assets = []
        liabilities = []
        equity = []
        total_assets = Decimal("0.00")
        total_liabilities = Decimal("0.00")
        total_equity = Decimal("0.00")

        for row in data:
            bal = Decimal(str(row.balance))
            if row.nature == "ASSET":
                assets.append(FinancialStatementItem(name=row.group_name, amount=float(bal)))
                total_assets += bal
            elif row.nature == "LIABILITY":
                # Liability usually has credit balance
                abs_bal = -bal
                liabilities.append(FinancialStatementItem(name=row.group_name, amount=float(abs_bal)))
                total_liabilities += abs_bal
            # TODO: Equity nature or special group for equity

        # P&L Net Profit should be added to Equity/Reserves
        pl = await self.get_profit_loss(company_id, fy_id)
        equity.append(FinancialStatementItem(name="Profit & Loss A/c", amount=pl.net_profit))
        total_equity += Decimal(str(pl.net_profit))

        return BalanceSheetResponse(
            assets=BalanceSheetSection(groups=assets, total=float(total_assets)),
            liabilities=BalanceSheetSection(groups=liabilities, total=float(total_liabilities)),
            equity=BalanceSheetSection(groups=equity, total=float(total_equity)),
            is_balanced=abs(total_assets - (total_liabilities + total_equity)) < 0.01
        )

    async def export_trial_balance_excel(self, company_id: uuid.UUID, fy_id: uuid.UUID) -> io.BytesIO:
        tb = await self.get_trial_balance(company_id, fy_id)

        wb = Workbook()
        ws = wb.active
        ws.title = "Trial Balance"

        # Headers
        headers = ["Ledger Name", "Opening", "Debit", "Credit", "Closing"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Data
        for item in tb.items:
            ws.append([
                item.ledger_name,
                item.opening_balance,
                item.debit_total,
                item.credit_total,
                item.closing_balance
            ])

        # Totals
        ws.append([])
        ws.append([
            "TOTALS",
            "",
            tb.total_debit,
            tb.total_credit,
            ""
        ])
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def get_warehouse_stock_report(self, company_id: uuid.UUID, warehouse_id: uuid.UUID) -> WarehouseStockResponse:
        warehouse = await self.db.get(Warehouse, warehouse_id)
        if not warehouse or warehouse.company_id != company_id:
            raise HTTPException(status_code=404, detail="Warehouse not found")

        rows = await get_warehouse_stock(self.db, company_id, warehouse_id)

        items = []
        total_val = Decimal("0.00")
        for row in rows:
            val = Decimal(str(row.quantity)) * Decimal(str(row.average_cost))
            items.append(WarehouseStockItem(
                item_id=row.item_id,
                item_name=row.item_name,
                quantity=float(row.quantity),
                average_cost=float(row.average_cost),
                value=float(val)
            ))
            total_val += val

        return WarehouseStockResponse(
            warehouse_name=warehouse.name,
            items=items,
            total_value=float(total_val)
        )

    async def get_transfer_history_report(self, company_id: uuid.UUID, fy_id: uuid.UUID) -> List[TransferReportItem]:
        rows = await get_transfer_history(self.db, company_id, fy_id)
        return [TransferReportItem(
            id=row.id,
            transfer_no=row.transfer_no,
            date=row.date,
            from_warehouse=row.from_warehouse,
            to_warehouse=row.to_warehouse,
            item_count=row.item_count,
            status=row.status
        ) for row in rows]
